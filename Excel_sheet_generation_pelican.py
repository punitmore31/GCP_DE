import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

# ==========================================
# CONFIGURATION
# ==========================================
workflow_map = {
    "KAIKU.wf_AR1_to_kaiku_stg": [
        "KAIKU_STG.STG_SOR_AR1_ACCOUNT",
        "KAIKU_STG.STG_SOR_AR1_CHARGE_GROUP_VIEW",
        "KAIKU_STG.STG_SOR_AR1_INVOICE",
    ],
    "KAIKU.wf_BASE_to_kaiku_stg": [
        "KAIKU_STG.STG_SVOC_BUS_FCT_CONTRACT_BASE_T",
        "KAIKU_STG.STG_SVOC_BUS_FCT_MOBILE_BASE_T",
    ],
    "KAIKU.wf_FIRA_to_kaiku_stg": [
        "KAIKU_STG.STG_FIRA_BUS_R_FCHD_KASITTELY",
        "KAIKU_STG.STG_FIRA_BUS_R_FCHD_TIKETTI",
        "KAIKU_STG.STG_FIRA_SA_CLIENT_IDENTITY",
    ],
    "KAIKU.wf_INVOICE_to_kaiku_stg": ["KAIKU_STG.STG_CARRY_OVER_INVOICES"],
    "KAIKU.wf_InitPartyIdLkp": [
        "KAIKU_STG.STG_SVOC_SOR_CONTRACT_LKP",
        "KAIKU_STG.STG_SVOC_SOR_CONTRACT_PARTY_LKP",
        "KAIKU_STG.STG_PRODUCT_TYPES",
        "KAIKU_STG.STG_SVOC_SOR_CNTR_PROD_LKP",
    ],
    "KAIKU.wf_InvoiceData_ToKaikuSTG": [
        "KAIKU_STG.STG_SVOC_SOR_SOR_PRODUCT",
        "KAIKU_STG.STG_SVOC_SOR_SOR_PARTY",
        "KAIKU_STG.KAIKU_PERIOD_FREEZE",
        "KAIKU_STG.ST_INVOICE_HEADER_ROW",
        "KAIKU_STG.STG_SVOC_SOR_SOR_INVOICE_DATA",
        "KAIKU_STG.STG_SVOC_SOR_SOR_PARTY_CONTRACT",
        "KAIKU_STG.STG_SVOC_SOR_SOR_CONTRACT",
    ],
    "KAIKU.wf_JPD_to_kaiku_stg": [
        "KAIKU_STG.STG_DIM_PARTY_SALESPERSON_JPD",
        "KAIKU_STG.STG_FL_JPD_HANDSETMODEL",
        "KAIKU_STG.STG_FL_JPD_HANDSETORDER",
        "KAIKU_STG.STG_FL_JPD_TICKET_CATEGORY",
        "KAIKU_STG.STG_FL_JPD_TICKETS",
        "KAIKU_STG.STG_FL_JPD_TITO_AGENT_ACTIONS",
        "KAIKU_STG.STG_SVOC_FL_JPD_DEVICEORDER",
        "KAIKU_STG.STG_SVOC_FL_JPD_GSMORDERS",
        "KAIKU_STG.STG_SVOC_FL_JPD_INTRATUNNUS",
        "KAIKU_STG.STG_TERRA_V_SOR_ELISA_JPD_MNP",
        "KAIKU_STG.STG_FL_JPD_TICKETS_CLOSED_HIST",
        "KAIKU_STG.STG_PROOMU_GSMORDERS",
        "KAIKU_STG.STG_PROOMU_GSM",
        "KAIKU_STG.STG_PROOMU_LOGICALSIM",
        "KAIKU_STG.STG_PROOMU_ORDER_GROUP_DELIVERY",
        "KAIKU_STG.STG_PROOMU_PHYSICALSIM",
        "KAIKU_STG.STG_PROOMU_PRODUCT",
        "KAIKU_STG.STG_PROOMU_PRODUCT_ORDER_GROUP",
        "KAIKU_STG.STG_PROOMU_PRODUCT_ORDER_ITEM",
        "KAIKU_STG.STG_PROOMU_CHANGE_SIM_ORDER",
        "KAIKU_STG.STG_FL_JPD_BILLINGITEMORDER",
        "KAIKU_STG.STG_DL_JPD_ORDERS",
        "KAIKU_STG.STG_DL_JPD_ORDERPARAMS",
    ],
    "KAIKU.wf_ODW_to_kaiku_stg": [
        "KAIKU_STG.STG_ODW_BUS4_BUS_TELLUS_RATED",
        "KAIKU_STG.STG_ODW_SOR2_SOR_MOBIN_CON_CONP",
        "KAIKU_STG.STG_ODW_BUS4_BUS_BASSET_RATED_INTERIM",
        "KAIKU_STG.LKP_MSISDN_NORMALIZER",
        "KAIKU_STG.STG_ODW_BUS4_BUS_BASSET_RATED",
    ],
    "KAIKU.wf_Prepaid": [
        "KAIKU_STG.STG_PREPAID_CRS_USAGE_MO",
        "KAIKU_STG.STG_PREPAID_CRS_USAGE_MT",
        "KAIKU_STG.STG_PREPAID_CRS_SERVICE_FEE",
        "KAIKU_STG.STG_PREPAID_CRS_ACCOUNT_COST",
        "KAIKU_STG.STG_PREPAID_CRS_PAY_VOU_REFILL",
    ],
    "KAIKU.wf_Rambo": [
        "KAIKU_STG.STG_OTSO_COSTS_FOR_KAIKU",
        "KAIKU_STG.STG_RAMBO_V_KAIKU_BTSDATA",
        "KAIKU_STG.STG_ODW_DM_MOBILE_TRAFFIC",
    ],
    "KAIKU.wf_Roaming": [
        "KAIKU_STG.STG_ODW_BUS_TAP_ROAMING_OUTBOUND",
        "KAIKU_STG.STG_ODW_BUS_TAP_ROAMING_INBOUND",
    ],
    "KAIKU.wf_SMS_IN_workaround": [
        "KAIKU_STG.tmp_SMS_IN_workaround",
        "KAIKU_STG.STG_ODW_BUS4_BUS_BASSET_RATED",
    ],
    "KAIKU.wf_TERRA_to_kaiku_stg": [
        "KAIKU_STG.STG_TERRA_BUS_DIM_UECC_AGENT",
        "KAIKU_STG.STG_TERRA_BUS_DIM_UECC_CALLTYPE",
        "KAIKU_STG.STG_TERRA_DIM_BILLING_ARRANGEMEN",
        "KAIKU_STG.STG_TERRA_DIM_CHARGE",
        "KAIKU_STG.STG_TERRA_DIM_CLASSIFICATION",
        "KAIKU_STG.STG_TERRA_DIM_CUSTOMER",
        "KAIKU_STG.STG_TERRA_DIM_SUBSCRIBER",
        "KAIKU_STG.STG_TERRA_DIM_UECC_AGENT_LOGIN",
        "KAIKU_STG.STG_TERRA_FACT_CREDIT_LOSS",
        "KAIKU_STG.STG_TERRA_FACT_UECC_KOSYSE",
    ],
    "KAIKU.wf_UsageIC": [
        "KAIKU_STG.STG_ODW_BUS_SMS_IC_TERM",
        "KAIKU_STG.STG_ODW_BUS_SMS_IC_ORIG",
    ],
    "KAIKU.wf_VAS": [
        "KAIKU_STG.STG_VAS_SMS",
        "KAIKU_STG.STG_VAS_MOBIILIMAKSAMINEN",
        "KAIKU_STG.STG_VAS_VOICE",
    ],
    "KAIKU.wf_device_to_kaiku_stg": ["KAIKU_STG.STG_SVOC_SOR_SOR_DEVICE_MODEL"],
    "KAIKU.wf_dynamo_kaiku_stg": [
        "KAIKU_STG.STG_DYNAMO_FACT_KAIKU",
        "KAIKU_STG.STG_SVOCSTG_FL_EMCE_BILLING_CARD",
        "KAIKU_STG.STG_SVOCSTG_FLEMCE_BUSINESS_PROD",
        "KAIKU_STG.STG_SVOCSTG_FLEMCE_NZ_COMPANY_V",
    ],
    "KAIKU.wf_enrich_sales_to_kaikustg": ["KAIKU_STG.STG_SVOC_STG_ST_EP_ENRICH"],
    "KAIKU.wf_fixed_base_to_kaiku_stg": ["KAIKU_STG.STG_SVOC_BUS_FCT_FIXED_BASE_T"],
    "KAIKU.wf_initKaikuMeta": ["KAIKU_STG.C_CONTROL_MASTER"],
    "KAIKU.wf_kaiku_datamart_report": [
        "KAIKU_STG.DL_KAIKU_PRODUCT_HIERARCHY",
        "KAIKU_STG.FL_KAIKU_PRODUCT_HIERARCHY",
        "KAIKU_DM.REP_QV_KAIKU_INTERIM",
        "KAIKU_DM.DIM_KAIKU_CUSTOMER",
        "KAIKU_DM.REP_QV_KAIKU",
        "KAIKU_DM.CTL_KAIKU_REP",
        "KAIKU_DM.TMP_DIM_KAIKU_CUSTOMER",
        "KAIKU_DM.DIM_KAIKU_CUSTOMER_HIST",
        "KAIKU_DM.TMP_DIM_KAIKU_CUSTOMER_DUPL",
    ],
    "KAIKU.wf_kaiku_telco_cloud": [
        "KAIKU_STG.FILE_PRESENCE_CHECK_VIDERA",
        "KAIKU_STG.STG_TELCOCLOUD_USAGE",
    ],
    "KAIKU.wf_mankeli_hierarchy_to_kaiku_stg": [
        "KAIKU_STG.STG_MANKELI_DIM_ACCT_HIERARCHY",
        "KAIKU_STG.STG_MANKELI_DIM_ORG_HIERARCHY",
        "KAIKU_STG.STG_MANKELI_DIM_VVFIELD_HIER",
    ],
    "KAIKU.wf_mankeli_to_kaiku_stg": [
        "KAIKU_STG.STG_MANKELI_ORDER_MASTER_ASSETS",
        "KAIKU_STG.STG_MANKELI_ST_ANLA",
        "KAIKU_STG.STG_SOR_ANLA",
        "KAIKU_STG.STG_MANKELI_FACT_REVENUE",
        "KAIKU_STG.STG_MANKELI_FACT_ORDER",
        "KAIKU_STG.STG_MANKELI_FACT_COPA_ALL",
        "KAIKU_STG.STG_MANKELI_FACT_SAP_LINE_ITEM",
        "KAIKU_STG.STG_MANKELI_FACT_SAP_ALL",
        "KAIKU_STG.STG_MANKELI_DIM_COMPANY_CODE",
        "KAIKU_STG.STG_MANKELI_DIM_ORDER_HIERARCHY",
        "KAIKU_STG.STG_MANKELI_DIM_CUST_HIERARCHY",
        "KAIKU_STG.STG_MANKELI_DIM_PROD_HIERARCHY",
    ],
    "KAIKU.wf_nirap_tables_to_kaiku_stg": [
        "KAIKU_STG.STG_NIRAP_A_KAIKU_FIXED",
        "KAIKU_STG.STG_NIRAP_A_KAIKU_MOBILE",
        "KAIKU_STG.STG_NIRAP_A_FTTH_FIBER",
        "KAIKU_STG.STG_NIRAP_A_KAIKU_CORE",
        "KAIKU_STG.STG_NIRAP_A_KAIKU_MOBILE_FIBER",
        "KAIKU_STG.STG_NIRAP_V_KAIKU_TRAFFIC",
    ],
    "KAIKU.wf_odw_busToKaikuSTG": [
        "KAIKU_STG.STG_DIM_CHARGECODE",
        "KAIKU_STG.STG_ODW_BUS4_BUS_TELLUS_HIST_RAW",
        "KAIKU_STG.STG_KAIKU_PROFIT_CENTER",
        "KAIKU_STG.STG_ODW_BUS4_BUS_TELLUS_HIST_F",
    ],
    "KAIKU.wf_sales_to_kaiku_stg": [
        "KAIKU_STG.STG_RM_EBOOK_EVENTS",
        "KAIKU_STG.STG_SVOC_STG_FL_KIRJA_SALES",
        "KAIKU_STG.STG_SVOC_STG_FL_KIRJA_CUSTOMER",
        "KAIKU_STG.STG_FCT_FIXED_SALES_T",
        "KAIKU_STG.STG_FCT_MOBILE_SALES_T",
        "KAIKU_STG.STG_SVOC_SOR_SOR_SALES",
        "KAIKU_STG.STG_SVOC_SOR_SOR_SALES_ITEM",
        "KAIKU_STG.STG_TERRA_DATA_MIPA_SOPIMUSTUOTE",
        "KAIKU_STG.STG_TERRA_DIM_MIPA_SOPIMUSHIER",
        "KAIKU_STG.STG_OE_HISTORICALORDER",
        "KAIKU_STG.STG_ODW_BUS4_DIM_BASSET_CUSTOMER",
        "KAIKU_STG.STG_SACC_STOCK_ITEM",
        "KAIKU_STG.STG_MOVING_AVG_PRICE",
        "KAIKU_STG.STG_YA_DEVICE_SALES",
        "KAIKU_STG.STG_FCT_CONTRACT_SALES_T",
        "KAIKU_STG.STG_SVOC_SOR_SOR_INVOICE",
    ],
    "KAIKU.wf_salesforce_kaiku_stg": [
        "KAIKU_STG.STG_SFDC_FL_OPPORTUNITIES",
        "KAIKU_STG.STG_SFDC_FL_EVENT",
        "KAIKU_STG.STG_SFDC_FL_BASE_MILESTONE1_PROJECT",
        "KAIKU_STG.STG_SFDC_FL_BASE_MILESTONE1_TASK",
        "KAIKU_STG.STG_SFDC_FL_BASE_MILESTONE1_TIME",
        "KAIKU_STG.STG_SFDC_V_ORDERS_SERVICE",
        "KAIKU_STG.STG_SFDC_FL_RESOURCE_REQUEST",
        "KAIKU_STG.STG_SFDC_FL_OPP_TEAMMEMBER",
        "KAIKU_STG.STG_SOR_OC_CONTACT_EVENT",
        "KAIKU_STG.STG_CTIREP_SOR_CASE",
        "KAIKU_STG.STG_SFDC_FL_ACCOUNT",
        "KAIKU_STG.STG_SFDC_FL_USER",
    ],
}

OUTPUT_DIR = r"C:\Users\punitkumar.more\Documents\Elisa\gcp_de\AUTOMATION\output\DOC2"


# ==========================================
# HELPER FUNCTIONS
# ==========================================
def get_safe_sheet_name(full_table_name, used_names):
    """
    Smart Sheet Naming Strategy:
    1. Clean invalid Excel chars.
    2. Try using the Full Name (Schema + Table).
    3. If > 31 chars, use ONLY the Table Name.
    4. If still > 31 chars, truncate the Table Name.
    5. Ensure Uniqueness.
    """
    # 1. Clean invalid characters
    invalid_chars = r"[]:*?\/\\"
    clean_name = full_table_name
    for char in invalid_chars:
        clean_name = clean_name.replace(char, "")

    # 2. Strategy A: Try Full Name (e.g. "KAIKU_STG.STG_TABLE")
    candidate_name = clean_name

    # 3. Strategy B: If too long, remove schema (e.g. "STG_TABLE")
    if len(candidate_name) > 31:
        if "." in candidate_name:
            candidate_name = candidate_name.split(".")[-1]

    # 4. Strategy C: If still too long, hard truncate
    if len(candidate_name) > 31:
        candidate_name = candidate_name[:31]

    # 5. Handle Duplicates
    counter = 1
    final_name = candidate_name

    while final_name in used_names:
        # Create space for suffix (e.g. "_1")
        suffix = f"_{counter}"
        trim_length = 31 - len(suffix)
        final_name = f"{candidate_name[:trim_length]}{suffix}"
        counter += 1

    used_names.add(final_name)
    return final_name


def apply_header_style(cell):
    """Applies bold font and light gray background to headers."""
    cell.font = Font(bold=True, size=12)
    # Light Blue/Gray fill
    # cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")


def create_section(ws, row_index, title, gap_after=2):
    """Writes a section header and returns the next available row index."""
    cell = ws.cell(row=row_index, column=1, value=title)
    apply_header_style(cell)
    return row_index + gap_after


# ==========================================
# MAIN LOGIC
# ==========================================
if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"Output Directory: {OUTPUT_DIR}\n")

    for wf_name, target_tables in workflow_map.items():
        print(f"Generating report for: {wf_name}...")

        wb = openpyxl.Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Track sheet names for this workbook to prevent duplicates
        used_sheet_names = set()

        for table_name in target_tables:
            # --- NAME FIX LOGIC ---
            safe_sheet_name = get_safe_sheet_name(table_name, used_sheet_names)

            print(f"  > Table: {table_name} -> Sheet: {safe_sheet_name}")

            ws = wb.create_sheet(title=safe_sheet_name)

            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 100

            current_row = 1

            # --- Section 1: Table Info ---
            # IMPORTANT: We write the ORIGINAL full table name in the cell
            ws.cell(row=current_row, column=1, value="Table Name :")
            apply_header_style(ws.cell(row=current_row, column=1))

            ws.cell(row=current_row, column=2, value=table_name)  # <--- Full Name Here
            # ws.cell(row=current_row, column=2).font = Font(bold=True)

            current_row += 2

            # --- Section 2: Incremental Validation ---
            current_row = create_section(
                ws, current_row, "Incremental Validation :", gap_after=2
            )

            # --- Section 3: Filter Used ---
            current_row = create_section(
                ws, current_row, "Filter used (if any) :", gap_after=3
            )

            # --- Section 4: Count Snippet ---
            current_row = create_section(
                ws, current_row, "Count Snippet :", gap_after=8
            )

            # --- Section 5: Data Snippet ---
            current_row = create_section(
                ws, current_row, "Data Snippet :", gap_after=10
            )

        # Save
        safe_filename = wf_name.replace("KAIKU.", "")  # Clean filename slightly
        file_name = f"{safe_filename}_Pelican.xlsx"
        full_path = os.path.join(OUTPUT_DIR, file_name)
        wb.save(full_path)
        print(f"✅ Saved: {full_path}\n")

    print("All reports generated successfully.")
